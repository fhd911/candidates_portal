# candidates/management/commands/import_users_roles.py
from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Optional, Sequence

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models.signals import post_save

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from candidates.models import Committee, MemberProfile

# ✅ signal الذي كان يسبب مشكلة إنشاء MemberProfile بدون committee
# عدّل المسار إذا كان عندك مختلف
from candidates.signals import ensure_profile  # noqa: F401

User = get_user_model()


# ======================================================
# Helpers
# ======================================================
def _norm(s: str) -> str:
    return "".join(str(s).strip().lower().split())


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _digits_only(v: Any) -> str:
    s = _s(v)
    return "".join(ch for ch in s if ch.isdigit())


def _bool(v: Any) -> bool:
    if v is None:
        return False
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"1", "true", "yes", "y", "نعم", "صح", "فعال", "مفعل", "نشط"}


def _has_field(model_cls, field_name: str) -> bool:
    try:
        model_cls._meta.get_field(field_name)
        return True
    except Exception:
        return False


def _get_header_map(ws: Worksheet) -> dict[str, int]:
    """
    returns: normalized_header -> 1-based column index
    """
    header_row = None
    for r in range(1, min(10, ws.max_row) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        non_empty = [v for v in vals if v not in (None, "", " ")]
        if len(non_empty) >= 2:
            header_row = r
            break
    if not header_row:
        raise CommandError("لم يتم العثور على صف عناوين (header) في الملف.")

    m: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = _norm(str(v))
        if key:
            m[key] = c
    m["__header_row__"] = header_row
    return m


def _col(header_map: dict[str, int], names: Sequence[str]) -> Optional[int]:
    for n in names:
        k = _norm(n)
        if k in header_map:
            return header_map[k]
    return None


def _get_cell(ws: Worksheet, row: int, col: Optional[int]) -> Any:
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def _role_choices_set() -> set[str]:
    try:
        return {str(k) for (k, _label) in MemberProfile._meta.get_field("role").choices}
    except Exception:
        return set()


def _committee_required_roles() -> set[str]:
    """
    إذا عندك ثابت MemberProfile.COMMITTEE_ROLES نستخدمه.
    غير ذلك: نفترض الشائع chair/member (وغيرها إن كانت موجودة في choices).
    """
    if hasattr(MemberProfile, "COMMITTEE_ROLES"):
        try:
            return set(getattr(MemberProfile, "COMMITTEE_ROLES"))
        except Exception:
            pass

    common = {"chair", "member", "committee_member", "committee_chair"}
    choices = _role_choices_set()
    return (common & choices) if choices else common


def _get_opportunity_model_and_keyfield() -> tuple[type, str]:
    """
    يستخرج موديل الفرصة من FK الموجود داخل Committee.opportunity
    ثم يختار أفضل حقل نصي متاح لإنشاء/جلب الفرصة (name -> title -> label).
    """
    try:
        f = Committee._meta.get_field("opportunity")
    except Exception as e:
        raise CommandError(
            "Model Committee لا يحتوي حقل opportunity. "
            "لكن قاعدة البيانات عندك تقول opportunity_id إجباري. راجع الموديل.\n"
            f"{e}"
        )

    opp_model = getattr(f, "related_model", None)
    if opp_model is None:
        raise CommandError("تعذر تحديد موديل الفرصة المرتبط بـ Committee.opportunity.")

    for cand in ("name", "title", "label"):
        if _has_field(opp_model, cand):
            return opp_model, cand

    raise CommandError(
        f"موديل الفرصة ({opp_model.__name__}) لا يحتوي حقول نصية name/title/label. "
        "أرسل لي الموديل وسأضبطه."
    )


def _create_or_get_opportunity(opp_model: type, key_field: str, opp_value: str):
    """
    get_or_create على موديل الفرصة باستخدام الحقل النصي المناسب
    """
    if not opp_value:
        raise CommandError("قيمة opportunity فارغة ولا يمكن إنشاء/جلب فرصة.")
    kwargs = {key_field: opp_value}
    return opp_model.objects.get_or_create(**kwargs)


def _mp_opportunity_is_fk() -> bool:
    if not _has_field(MemberProfile, "opportunity"):
        return False
    f = MemberProfile._meta.get_field("opportunity")
    return bool(getattr(f, "is_relation", False))


# ======================================================
# Data
# ======================================================
@dataclass
class RowData:
    national_id: str
    full_name: str
    role: str
    is_active: bool
    opportunity: str
    committee_name: str


# ======================================================
# Command
# ======================================================
class Command(BaseCommand):
    help = "Import users + roles + committee membership from an Excel file."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to .xlsx file")
        parser.add_argument("--dry-run", action="store_true", help="Do not write to DB")

    def handle(self, *args, **options):
        path: str = options["xlsx_path"]
        dry: bool = bool(options["dry_run"])

        try:
            wb = load_workbook(filename=path, data_only=True)
        except Exception as e:
            raise CommandError(f"تعذر فتح الملف: {path}\n{e}")

        ws = wb.active
        header_map = _get_header_map(ws)
        header_row = int(header_map["__header_row__"])

        # Columns (Arabic/English)
        c_nid = _col(header_map, ["national_id", "nid", "identity", "id", "السجل", "السجل_المدني", "الهوية", "رقم_الهوية"])
        c_name = _col(header_map, ["full_name", "name", "الاسم", "الاسم_كامل", "الاسم_الكامل"])
        c_role = _col(header_map, ["role", "الدور", "الصفة", "صلاحية", "المسمى"])
        c_active = _col(header_map, ["active", "is_active", "enabled", "تفعيل", "فعال", "نشط", "مفعل"])
        c_opp = _col(header_map, ["opportunity", "opp", "الفرصة", "فرصة"])
        c_committee = _col(header_map, ["committee", "committee_name", "اللجنة", "لجنة", "committee"])

        if not c_nid:
            raise CommandError("لا يوجد عمود للسجل/الهوية. تأكد من وجود عمود مثل: السجل المدني / national_id")

        committee_required_roles = _committee_required_roles()
        valid_roles = _role_choices_set()

        # Opportunity model inferred from Committee.opportunity FK
        opp_model, opp_key_field = _get_opportunity_model_and_keyfield()

        self.stdout.write(self.style.MIGRATE_HEADING(f"Reading: {path}"))
        self.stdout.write(f"Sheet: {ws.title} | header row: {header_row}")
        self.stdout.write(f"Opportunity model: {opp_model.__name__} (key field: {opp_key_field})")
        if dry:
            self.stdout.write(self.style.WARNING("DRY RUN: لن يتم حفظ أي شيء في قاعدة البيانات (Dry-run حقيقي)."))

        mp_opp_fk = _mp_opportunity_is_fk()

        # Collect rows
        rows: list[RowData] = []
        for r in range(header_row + 1, ws.max_row + 1):
            nid = _digits_only(_get_cell(ws, r, c_nid))
            if not nid:
                continue

            full_name = _s(_get_cell(ws, r, c_name))
            role = _s(_get_cell(ws, r, c_role))

            # fallback role (لا نخمن roles جديدة؛ نستخدم أول choice إن لم يوجد)
            if not role:
                if valid_roles:
                    role = next(iter(valid_roles))
                else:
                    role = "member"

            is_active = _bool(_get_cell(ws, r, c_active))
            opportunity = _s(_get_cell(ws, r, c_opp))
            committee_name = _s(_get_cell(ws, r, c_committee))

            if valid_roles and role not in valid_roles:
                raise CommandError(f"Row {r}: role='{role}' غير موجود ضمن choices في MemberProfile.role")

            # إذا دور لجنة لازم تكون اللجنة موجودة + لازم يكون opportunity موجود لأن Committee يتطلبه
            if role in committee_required_roles:
                if not committee_name:
                    raise CommandError(f"Row {r}: role='{role}' يتطلب لجنة (committee) لكن العمود فارغ.")
                if not opportunity:
                    raise CommandError(
                        f"Row {r}: committee='{committee_name}' تتطلب opportunity لكن عمود opportunity/الفرصة فارغ."
                    )

            rows.append(
                RowData(
                    national_id=nid,
                    full_name=full_name,
                    role=role,
                    is_active=is_active,
                    opportunity=opportunity,
                    committee_name=committee_name,
                )
            )

        if not rows:
            self.stdout.write(self.style.WARNING("لا توجد صفوف صالحة للاستيراد."))
            return

        # Counters
        created_users = 0
        existing_users = 0
        created_profiles = 0
        updated_profiles = 0
        created_committees = 0
        created_opps = 0

        # ✅ افصل signal أثناء الاستيراد الحقيقي (حتى لا ينشئ profile مخالف للقيود)
        disconnected = False
        if not dry:
            try:
                post_save.disconnect(ensure_profile, sender=User)
                disconnected = True
            except Exception:
                disconnected = False

        try:
            if dry:
                # ======================================================
                # DRY RUN (NO DB)
                # ======================================================
                for i, rd in enumerate(rows, start=1):
                    if rd.opportunity:
                        self.stdout.write(f"[DRY] Would get_or_create {opp_model.__name__}({opp_key_field}='{rd.opportunity}')")

                    if rd.committee_name:
                        if not rd.opportunity:
                            raise CommandError(f"Row {i}: committee='{rd.committee_name}' يتطلب opportunity لكن قيمته فارغة.")
                        self.stdout.write(f"[DRY] Would get_or_create Committee(name='{rd.committee_name}', opportunity='{rd.opportunity}')")

                    self.stdout.write(f"[DRY] Would get_or_create User(username={rd.national_id})")
                    self.stdout.write(
                        f"[DRY] Would update_or_create MemberProfile(user={rd.national_id}, "
                        f"role={rd.role}, active={rd.is_active}, committee={rd.committee_name or '-'}, "
                        f"opportunity={'FK instance' if mp_opp_fk else (rd.opportunity or '-')})"
                    )
                    self.stdout.write(
                        f"Row {i}: {rd.national_id} role={rd.role} active={rd.is_active} "
                        f"opp={rd.opportunity or '-'} com={rd.committee_name or '-'}"
                    )

                self.stdout.write(self.style.SUCCESS("DRY RUN Done ✅ (no database writes)"))
                return

            # ======================================================
            # REAL RUN (DB writes)
            # ======================================================
            with transaction.atomic():
                for i, rd in enumerate(rows, start=1):
                    # Opportunity instance (if provided)
                    opp_obj = None
                    if rd.opportunity:
                        opp_obj, opp_created = _create_or_get_opportunity(opp_model, opp_key_field, rd.opportunity)
                        if opp_created:
                            created_opps += 1

                    # Committee (requires opportunity instance)
                    committee_obj = None
                    if rd.committee_name:
                        if not opp_obj:
                            raise CommandError(
                                f"Row {i}: committee='{rd.committee_name}' يتطلب opportunity لكن لم يتم تحديدها."
                            )
                        committee_obj, com_created = Committee.objects.get_or_create(
                            name=rd.committee_name,
                            opportunity=opp_obj,
                        )
                        if com_created:
                            created_committees += 1

                    # User
                    u, u_created = User.objects.get_or_create(username=rd.national_id)
                    if u_created:
                        created_users += 1
                        if hasattr(u, "set_unusable_password"):
                            u.set_unusable_password()
                    else:
                        existing_users += 1

                    if rd.full_name:
                        u.first_name = rd.full_name[:30]
                    u.save()

                    # Profile
                    defaults: dict[str, Any] = {
                        "role": rd.role,
                        "committee": committee_obj,
                        "is_active": rd.is_active,
                    }

                    # opportunity in MemberProfile: FK -> instance, else -> string
                    if _has_field(MemberProfile, "opportunity"):
                        mp_f = MemberProfile._meta.get_field("opportunity")
                        if getattr(mp_f, "is_relation", False):
                            defaults["opportunity"] = opp_obj
                        else:
                            defaults["opportunity"] = rd.opportunity

                    if _has_field(MemberProfile, "full_name") and rd.full_name:
                        defaults["full_name"] = rd.full_name

                    prof, p_created = MemberProfile.objects.update_or_create(user=u, defaults=defaults)
                    if p_created:
                        created_profiles += 1
                    else:
                        updated_profiles += 1

                    self.stdout.write(
                        f"Row {i}: {rd.national_id} role={rd.role} active={rd.is_active} "
                        f"opp={rd.opportunity or '-'} com={rd.committee_name or '-'}"
                    )

        finally:
            if not dry and disconnected:
                try:
                    post_save.connect(ensure_profile, sender=User)
                except Exception:
                    pass

        self.stdout.write(self.style.SUCCESS("Done ✅"))
        self.stdout.write(
            f"Opportunities created: {created_opps}\n"
            f"Committees created: {created_committees}\n"
            f"Users created: {created_users} | existing users touched: {existing_users}\n"
            f"Profiles created: {created_profiles} | profiles updated: {updated_profiles}\n"
            f"Signal disconnected during import: {disconnected}"
        )
