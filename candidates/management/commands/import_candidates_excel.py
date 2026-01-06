# candidates/management/commands/import_candidates_excel.py
from __future__ import annotations

from pathlib import Path
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from candidates.models import Candidate, Opportunity, Committee


AR_MAP = {
    "اسم المتقدم": "full_name",
    "السجل المدني": "national_id",
    "رقم الجوال": "mobile",
    "التخصص": "specialization",
    "الرتبة الوظيفية": "rank",
    "العمل الحالي": "current_work",
    "تاريخ المباشرة (هجري)": "start_date_hijri",
    "مدرسة المتقدم": "school",
    "قطاع المتقدم": "sector",
    "الوظيفة المتقدم عليها": "applied_position",
    "مدرسة الفرصة": "opportunity_school",
    "قطاع الفرصة": "opportunity_sector",
    "سبق العمل في الإدارة المدرسية": "admin_exp",
    "سنوات عمل مدير": "years_director",
    "سنوات عمل وكيل": "years_deputy",
    "رابط السيرة الذاتية": "cv_url",
}


def clean_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def clean_int(v) -> int:
    s = clean_str(v)
    if not s:
        return 0
    try:
        return int(float(s))
    except Exception:
        return 0


def normalize_mobile(v: str) -> str:
    digits = re.sub(r"\D+", "", clean_str(v))
    # أحيانًا يجي 9665xxxxxxxx — نخليه كما هو (أرقام فقط)
    return digits


class Command(BaseCommand):
    help = "Import candidates from Arabic Excel headers into DB."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to .xlsx")
        parser.add_argument("--opportunity", required=True, type=str, help="Opportunity name (e.g. وكيل مدرسة 1447)")
        parser.add_argument("--committee", type=str, default="", help="Committee name (optional) e.g. اللجنة 1")
        parser.add_argument("--assign", action="store_true", help="Assign imported candidates to the committee")

    @transaction.atomic
    def handle(self, *args, **opts):
        xlsx_path = Path(opts["xlsx_path"]).expanduser().resolve()
        if not xlsx_path.exists():
            raise CommandError(f"File not found: {xlsx_path}")

        opp_name = clean_str(opts["opportunity"])
        if not opp_name:
            raise CommandError("--opportunity is required and must not be empty.")

        opportunity, _ = Opportunity.objects.get_or_create(name=opp_name, defaults={"is_active": True})

        committee_name = clean_str(opts["committee"])
        assign = bool(opts["assign"])

        committee = None
        if committee_name:
            committee, _ = Committee.objects.get_or_create(
                opportunity=opportunity,
                name=committee_name,
                defaults={"is_open": True},
            )

        wb = load_workbook(xlsx_path)
        ws = wb.active

        # header row
        headers = [clean_str(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        idx = {h: i for i, h in enumerate(headers)}

        missing = [h for h in AR_MAP.keys() if h not in idx]
        if missing:
            raise CommandError(f"Missing columns in Excel: {missing}")

        created = 0
        updated = 0
        skipped = 0

        for r in range(2, ws.max_row + 1):
            row = {h: ws.cell(r, idx[h] + 1).value for h in AR_MAP.keys()}

            national_id = clean_str(row["السجل المدني"])
            full_name = clean_str(row["اسم المتقدم"])
            if not national_id or not full_name:
                skipped += 1
                continue

            defaults = {
                "full_name": full_name,
                "mobile": normalize_mobile(row["رقم الجوال"]),
                "specialization": clean_str(row["التخصص"]),
                "rank": clean_str(row["الرتبة الوظيفية"]),
                "current_work": clean_str(row["العمل الحالي"]),
                "start_date_hijri": clean_str(row["تاريخ المباشرة (هجري)"]),
                "school": clean_str(row["مدرسة المتقدم"]),
                "sector": clean_str(row["قطاع المتقدم"]),
                "applied_position": clean_str(row["الوظيفة المتقدم عليها"]),
                "opportunity_school": clean_str(row["مدرسة الفرصة"]),
                "opportunity_sector": clean_str(row["قطاع الفرصة"]),
                "admin_exp": clean_str(row["سبق العمل في الإدارة المدرسية"]),
                "years_director": clean_int(row["سنوات عمل مدير"]),
                "years_deputy": clean_int(row["سنوات عمل وكيل"]),
                "cv_url": clean_str(row["رابط السيرة الذاتية"]),
            }

            obj, is_created = Candidate.objects.update_or_create(
                opportunity=opportunity,
                national_id=national_id,
                defaults=defaults,
            )

            if committee and assign:
                obj.assigned_committee = committee
                obj.save(update_fields=["assigned_committee"])

            created += 1 if is_created else 0
            updated += 0 if is_created else 1

        self.stdout.write(self.style.SUCCESS(
            f"Import done: created={created}, updated={updated}, skipped={skipped}, opportunity='{opportunity.name}'"
            + (f", committee='{committee.name}', assigned={assign}" if committee else "")
        ))
